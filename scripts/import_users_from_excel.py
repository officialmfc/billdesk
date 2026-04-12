#!/usr/bin/env python3
"""Bulk import public.users rows from an Excel workbook.

This script intentionally does not create Supabase auth users. It parses a
spreadsheet, normalizes the rows, and sends them to the auth hub's service-role
RPC for insertion into public.users.
"""

from __future__ import annotations

import argparse
import json
import os
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from supabase import create_client


REQUIRED_HEADERS = ["full_name", "phone", "business_name", "user_type", "default_role"]
OPTIONAL_HEADERS = ["address_json", "profile_photo_url"]
SUPPORTED_HEADERS = REQUIRED_HEADERS + OPTIONAL_HEADERS


@dataclass(slots=True)
class ImportRow:
    full_name: str
    phone: str
    business_name: str | None
    user_type: str
    default_role: str
    address_json: Any | None
    profile_photo_url: str | None


def normalize_header(value: Any) -> str:
    return str(value or "").strip().lower().replace(" ", "_")


def normalize_phone(value: Any) -> str:
    raw = str(value or "").strip()
    if not raw:
        return ""
    keep = []
    for char in raw:
        if char.isdigit() or char == "+":
            keep.append(char)
    return "".join(keep)


def normalize_text(value: Any) -> str:
    return str(value or "").strip()


def parse_address_json(value: Any) -> Any | None:
    if value is None:
        return None
    if isinstance(value, (dict, list)):
        return value
    text = str(value).strip()
    if not text:
        return None
    try:
        return json.loads(text)
    except json.JSONDecodeError as error:
        raise ValueError(f"address_json must be valid JSON: {error}") from error


def load_rows_from_workbook(path: Path, sheet_name: str | None) -> list[ImportRow]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]

    rows = worksheet.iter_rows(values_only=True)
    try:
        raw_headers = next(rows)
    except StopIteration as error:
        raise ValueError("The workbook does not contain any rows.") from error

    header_map = {normalize_header(header): index for index, header in enumerate(raw_headers)}
    missing_headers = [header for header in REQUIRED_HEADERS if header not in header_map]
    if missing_headers:
        raise ValueError(f"Missing required headers: {', '.join(missing_headers)}")

    unknown_headers = [
        normalize_header(header)
        for header in raw_headers
        if normalize_header(header) and normalize_header(header) not in SUPPORTED_HEADERS
    ]
    if unknown_headers:
        print(
            f"Warning: ignoring unknown columns: {', '.join(sorted(set(unknown_headers)))}",
            file=sys.stderr,
        )

    parsed_rows: list[ImportRow] = []
    for row_index, row in enumerate(rows, start=2):
        full_name = normalize_text(row[header_map["full_name"]] if header_map["full_name"] < len(row) else "")
        phone = normalize_phone(row[header_map["phone"]] if header_map["phone"] < len(row) else "")
        business_name = normalize_text(row[header_map["business_name"]] if header_map["business_name"] < len(row) else "")
        user_type = normalize_text(row[header_map["user_type"]] if header_map["user_type"] < len(row) else "").lower()
        default_role = normalize_text(row[header_map["default_role"]] if header_map["default_role"] < len(row) else "").lower()
        address_json = parse_address_json(
            row[header_map["address_json"]] if "address_json" in header_map and header_map["address_json"] < len(row) else None
        )
        profile_photo_url = normalize_text(
            row[header_map["profile_photo_url"]]
            if "profile_photo_url" in header_map and header_map["profile_photo_url"] < len(row)
            else ""
        )

        if not any([full_name, phone, business_name, user_type, default_role, address_json, profile_photo_url]):
            continue

        parsed_rows.append(
            ImportRow(
                full_name=full_name,
                phone=phone,
                business_name=business_name or None,
                user_type=user_type,
                default_role=default_role,
                address_json=address_json,
                profile_photo_url=profile_photo_url or None,
            )
        )

    return parsed_rows


def validate_rows(rows: Iterable[ImportRow]) -> None:
    allowed_user_types = {"vendor", "business"}
    allowed_roles = {"buyer", "seller"}
    for index, row in enumerate(rows, start=1):
        if not row.full_name:
            raise ValueError(f"Row {index}: full_name is required.")
        if not row.phone:
            raise ValueError(f"Row {index}: phone is required.")
        if row.user_type not in allowed_user_types:
            raise ValueError(f"Row {index}: user_type must be vendor or business.")
        if row.default_role not in allowed_roles:
            raise ValueError(f"Row {index}: default_role must be buyer or seller.")


def to_rpc_payload(rows: Iterable[ImportRow]) -> list[dict[str, Any]]:
    payload: list[dict[str, Any]] = []
    for row in rows:
        item: dict[str, Any] = {
            "full_name": row.full_name,
            "phone": row.phone,
            "business_name": row.business_name,
            "user_type": row.user_type,
            "default_role": row.default_role,
            "profile_photo_url": row.profile_photo_url,
        }
        if row.address_json is not None:
            item["address_json"] = row.address_json
        payload.append(item)
    return payload


def build_summary(result: dict[str, Any]) -> str:
    lines = [
        f"Dry run: {bool(result.get('dry_run'))}",
        f"Total rows: {result.get('total', 0)}",
        f"Created: {result.get('created', 0)}",
        f"Skipped: {result.get('skipped', 0)}",
        f"Failed: {result.get('failed', 0)}",
    ]

    for section_name in ("created_rows", "skipped_rows", "failed_rows"):
        rows = result.get(section_name) or []
        if not rows:
            continue
        lines.append("")
        lines.append(section_name.replace("_", " ").title() + ":")
        for row in rows:
            row_number = row.get("row_number", "?")
            status = row.get("status", "unknown")
            reason = row.get("reason")
            label = row.get("full_name") or row.get("phone") or "row"
            suffix = f" - {reason}" if reason else ""
            lines.append(f"  - Row {row_number} [{status}] {label}{suffix}")

    return "\n".join(lines)


def main() -> int:
    parser = argparse.ArgumentParser(description="Bulk import users from an Excel workbook.")
    parser.add_argument("--file", required=True, help="Path to the Excel workbook.")
    parser.add_argument("--sheet", help="Optional worksheet name.")
    parser.add_argument("--dry-run", action="store_true", help="Validate and summarize without writing rows.")
    parser.add_argument(
        "--json-output",
        help="Write the RPC summary JSON to this path after the import completes.",
    )
    parser.add_argument(
        "--print-json",
        action="store_true",
        help="Print the RPC summary JSON to stdout instead of the human summary.",
    )
    args = parser.parse_args()

    supabase_url = os.getenv("SUPABASE_URL", "").strip() or os.getenv("NEXT_PUBLIC_SUPABASE_URL", "").strip()
    service_role_key = os.getenv("SUPABASE_SERVICE_ROLE_KEY", "").strip()
    if not supabase_url:
        print("Missing SUPABASE_URL or NEXT_PUBLIC_SUPABASE_URL.", file=sys.stderr)
        return 1
    if not service_role_key:
        print("Missing SUPABASE_SERVICE_ROLE_KEY.", file=sys.stderr)
        return 1

    workbook_path = Path(args.file).expanduser().resolve()
    if not workbook_path.exists():
        print(f"Workbook not found: {workbook_path}", file=sys.stderr)
        return 1

    try:
        rows = load_rows_from_workbook(workbook_path, args.sheet)
        validate_rows(rows)
    except Exception as error:
        print(f"Failed to parse workbook: {error}", file=sys.stderr)
        return 1

    client = create_client(supabase_url, service_role_key)
    payload = to_rpc_payload(rows)

    try:
        response = client.rpc(
            "bulk_create_users_from_import",
            {
                "p_rows": payload,
                "p_dry_run": args.dry_run,
            },
        ).execute()
    except Exception as error:
        print(f"RPC call failed: {error}", file=sys.stderr)
        return 1

    result = dict(response.data or {})

    if args.json_output:
        output_path = Path(args.json_output).expanduser().resolve()
        output_path.write_text(json.dumps(result, indent=2, sort_keys=True), encoding="utf-8")
        print(f"Wrote JSON summary to {output_path}")

    if args.print_json:
        print(json.dumps(result, indent=2, sort_keys=True))
    else:
        print(build_summary(result))

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
